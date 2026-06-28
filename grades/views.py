import io
import os
from datetime import datetime
from decimal import Decimal

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib.colors import HexColor, white, black
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    KeepTogether, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.db.models import Q

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from openpyxl import load_workbook

from .models import Term, Grade, GradeHistory, StudentAverage
from .serializers import TermSerializer, GradeSerializer, GradeHistorySerializer, StudentAverageSerializer
from accounts.permissions import CanManageGrades

from students.models import Student
from subjects.models import TeacherSubject

PRIMARY = HexColor('#1e40af')
SECONDARY = HexColor('#3b82f6')
LIGHT = HexColor('#eff6ff')
HEADER_BG = HexColor('#1e40af')
HEADER_FG = white
ROW_ALT = HexColor('#f8fafc')
GREEN = HexColor('#16a34a')
RED = HexColor('#dc2626')
ORANGE = HexColor('#ea580c')
GRAY = HexColor('#6b7280')
DARK = HexColor('#1f2937')

SCHOOL_NAME = "Ecole Privee Excellence"
SCHOOL_ADDRESS = "Conakry, Republique de Guinee"

def get_appreciation(avg):
    try:
        a = float(avg)
    except (TypeError, ValueError):
        return '--'
    if a >= 16:
        return 'Tres bien'
    if a >= 14:
        return 'Bien'
    if a >= 12:
        return 'Assez bien'
    if a >= 10:
        return 'Passable'
    return 'Insuffisant'

class TermViewSet(viewsets.ModelViewSet):
    queryset = Term.objects.all()
    serializer_class = TermSerializer
    permission_classes = [IsAuthenticated]

class GradeViewSet(viewsets.ModelViewSet):
    queryset = Grade.objects.all()
    serializer_class = GradeSerializer
    permission_classes = [IsAuthenticated, CanManageGrades]

    def perform_create(self, serializer):
        grade = serializer.save()
        self._log_changes(grade, created=True)

    def perform_update(self, serializer):
        old_instance = self.get_object()
        old_values = {
            'homework1': old_instance.homework1,
            'homework2': old_instance.homework2,
            'composition': old_instance.composition,
        }
        grade = serializer.save()
        for field, old_val in old_values.items():
            new_val = getattr(grade, field)
            if old_val != new_val:
                GradeHistory.objects.create(
                    grade=grade,
                    user=self.request.user,
                    field_name=field,
                    old_value=str(old_val) if old_val is not None else None,
                    new_value=str(new_val) if new_val is not None else None,
                )

    def perform_destroy(self, instance):
        GradeHistory.objects.create(
            grade=instance,
            user=self.request.user,
            field_name='deleted',
            old_value=str(instance),
            new_value=None,
        )
        instance.delete()

    def _log_changes(self, grade, created=False):
        if created:
            for field in ['homework1', 'homework2', 'composition']:
                val = getattr(grade, field)
                if val is not None:
                    GradeHistory.objects.create(
                        grade=grade,
                        user=self.request.user,
                        field_name=field,
                        old_value=None,
                        new_value=str(val),
                    )

    def get_queryset(self):
        queryset = Grade.objects.select_related('student', 'teacher_subject', 'term')
        student_id = self.request.query_params.get('student_id', None)
        class_id = self.request.query_params.get('class_id', None)
        subject_id = self.request.query_params.get('subject_id', None)
        term_id = self.request.query_params.get('term_id', None)
        teacher_id = self.request.query_params.get('teacher_id', None)

        if student_id:
            queryset = queryset.filter(student_id=student_id)
        if class_id:
            queryset = queryset.filter(student__class_assigned_id=class_id)
        if subject_id:
            queryset = queryset.filter(teacher_subject__subject_id=subject_id)
        if term_id:
            queryset = queryset.filter(term_id=term_id)
        if teacher_id:
            queryset = queryset.filter(teacher_subject__teacher_id=teacher_id)
        return queryset

    @action(detail=False, methods=['post'])
    def initialize(self, request):
        class_id = request.data.get('class_id')
        subject_id = request.data.get('subject_id')
        term_id = request.data.get('term_id')

        if not all([class_id, subject_id, term_id]):
            return Response({'error': 'class_id, subject_id, term_id sont requis'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            term = Term.objects.get(id=term_id)
        except Term.DoesNotExist:
            return Response({'error': 'Trimestre introuvable'}, status=status.HTTP_404_NOT_FOUND)

        from subjects.models import Subject
        from subjects.models import TeacherSubject
        from teachers.models import Teacher

        try:
            subject = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            return Response({'error': 'Matière introuvable'}, status=status.HTTP_404_NOT_FOUND)

        students = Student.objects.filter(class_assigned_id=class_id)
        if not students.exists():
            return Response({'error': 'Aucun élève dans cette classe'}, status=status.HTTP_400_BAD_REQUEST)

        teacher_subject = TeacherSubject.objects.filter(
            subject=subject,
            class_assigned_id=class_id
        ).first()

        if not teacher_subject:
            teacher = subject.teacher
            if not teacher:
                teacher = Teacher.objects.filter(is_active=True).first()
            if not teacher:
                return Response({
                    'error': 'Aucun enseignant trouvé dans le système. Créez d\'abord un enseignant.'
                }, status=status.HTTP_400_BAD_REQUEST)
            teacher_subject, _ = TeacherSubject.objects.get_or_create(
                teacher=teacher,
                subject=subject,
                class_assigned_id=class_id,
                defaults={'academic_year': term.academic_year}
            )

        created = 0
        for student in students:
            _, was_created = Grade.objects.get_or_create(
                student=student,
                teacher_subject=teacher_subject,
                term=term,
                defaults={'homework1': None, 'homework2': None, 'composition': None}
            )
            if was_created:
                created += 1

        return Response({
            'success': True,
            'message': f'{created} note(s) créée(s) sur {students.count()} élève(s)',
            'created': created,
            'total': students.count(),
        })

    @action(detail=False, methods=['post'])
    def recalculate_averages(self, request):
        student_id = request.data.get('student_id')
        term_id = request.data.get('term_id')

        if not term_id:
            return Response({'error': 'term_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        term = Term.objects.get(id=term_id)

        if student_id:
            students = Student.objects.filter(id=student_id)
        else:
            students = Student.objects.filter(status='active')

        for student in students:
            StudentAverage.calculate_average(student, term)

        StudentAverage.update_rankings(term)

        return Response({'status': 'Averages recalculated'})

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        if 'file' not in request.FILES:
            return Response({'error': 'Aucun fichier fourni'}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        if not file.name.endswith('.xlsx') and not file.name.endswith('.xls'):
            return Response({'error': 'Format invalide. Utilisez un fichier .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

        term_id = request.POST.get('term_id')
        class_id = request.POST.get('class_id')
        subject_id = request.POST.get('subject_id')

        if not all([term_id, class_id, subject_id]):
            return Response({
                'error': 'Parametres manquants: term_id, class_id, subject_id sont requis'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            term = Term.objects.get(id=term_id)
        except Term.DoesNotExist:
            return Response({'error': 'Trimestre introuvable'}, status=status.HTTP_404_NOT_FOUND)

        try:
            from subjects.models import Subject
            subject = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            return Response({'error': 'Matiere introuvable'}, status=status.HTTP_404_NOT_FOUND)

        students_in_class = Student.objects.filter(class_assigned_id=class_id)
        if not students_in_class.exists():
            return Response({'error': 'Aucun eleve dans cette classe'}, status=status.HTTP_400_BAD_REQUEST)

        student_map = {}
        for s in students_in_class:
            student_map[s.full_name.lower().strip()] = s
            student_map[s.matricule.lower().strip()] = s
            student_map[f"{s.last_name.lower()} {s.first_name.lower()}"] = s

        try:
            wb = load_workbook(file)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Erreur de lecture du fichier: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        success_count = 0
        error_list = []
        created = 0
        updated = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue

            try:
                identifier = str(row[0]).strip() if row[0] else None
                if not identifier:
                    error_list.append({
                        'row': row_idx,
                        'error': 'Nom/Matricule vide',
                        'data': str(row[:5])
                    })
                    continue

                student = student_map.get(identifier.lower())
                if not student:
                    error_list.append({
                        'row': row_idx,
                        'error': f'Eleve non trouve: {identifier}',
                        'data': str(row[:5])
                    })
                    continue

                def parse_note(val):
                    if val is None or val == '':
                        return None
                    try:
                        v = float(val)
                        if v < 0:
                            return None
                        return v
                    except (ValueError, TypeError):
                        return None

                homework1 = parse_note(row[1]) if len(row) > 1 else None
                homework2 = parse_note(row[2]) if len(row) > 2 else None
                composition = parse_note(row[3]) if len(row) > 3 else None

                if all(n is None for n in [homework1, homework2, composition]):
                    error_list.append({
                        'row': row_idx,
                        'error': 'Aucune note valide',
                        'data': str(row[:5])
                    })
                    continue

                from subjects.models import TeacherSubject
                teacher_subject = TeacherSubject.objects.filter(
                    subject=subject,
                    class_assigned_id=class_id
                ).first()

                if not teacher_subject:
                    error_list.append({
                        'row': row_idx,
                        'error': f'Enseignant non trouve pour {subject.name}',
                        'data': str(row[:5])
                    })
                    continue

                grade, was_created = Grade.objects.update_or_create(
                    student=student,
                    teacher_subject=teacher_subject,
                    term=term,
                    defaults={
                        'homework1': homework1,
                        'homework2': homework2,
                        'composition': composition,
                    }
                )

                if was_created:
                    created += 1
                else:
                    updated += 1
                success_count += 1

            except Exception as e:
                error_list.append({
                    'row': row_idx,
                    'error': f'Erreur: {str(e)}',
                    'data': str(row[:5])
                })

        return Response({
            'success': True,
            'message': f'{success_count} notes traitees ({created} creees, {updated} mises a jour)',
            'created': created,
            'updated': updated,
            'errors': error_list,
            'total_errors': len(error_list),
        })

    @action(detail=False, methods=['get'])
    def history(self, request):
        grade_id = request.query_params.get('grade_id', None)
        student_id = request.query_params.get('student_id', None)
        subject_id = request.query_params.get('subject_id', None)
        class_id = request.query_params.get('class_id', None)

        queryset = GradeHistory.objects.select_related('grade', 'user')
        if grade_id:
            queryset = queryset.filter(grade_id=grade_id)
        if student_id:
            queryset = queryset.filter(grade__student_id=student_id)
        if subject_id:
            queryset = queryset.filter(grade__teacher_subject__subject_id=subject_id)
        if class_id:
            queryset = queryset.filter(grade__student__class_assigned_id=class_id)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = GradeHistorySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = GradeHistorySerializer(queryset, many=True)
        return Response(serializer.data)


class GradeHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = GradeHistory.objects.select_related('grade', 'user').all()
    serializer_class = GradeHistorySerializer
    permission_classes = [IsAuthenticated, CanManageGrades]

    def get_queryset(self):
        queryset = self.queryset
        grade_id = self.request.query_params.get('grade_id', None)
        student_id = self.request.query_params.get('student_id', None)
        subject_id = self.request.query_params.get('subject_id', None)
        class_id = self.request.query_params.get('class_id', None)
        if grade_id:
            queryset = queryset.filter(grade_id=grade_id)
        if student_id:
            queryset = queryset.filter(grade__student_id=student_id)
        if subject_id:
            queryset = queryset.filter(grade__teacher_subject__subject_id=subject_id)
        if class_id:
            queryset = queryset.filter(grade__student__class_assigned_id=class_id)
        return queryset


class StudentAverageViewSet(viewsets.ModelViewSet):
    queryset = StudentAverage.objects.all()
    serializer_class = StudentAverageSerializer
    permission_classes = [IsAuthenticated, CanManageGrades]

    def get_queryset(self):
        queryset = StudentAverage.objects.all()
        term_id = self.request.query_params.get('term_id', None)
        if term_id:
            queryset = queryset.filter(term_id=term_id).order_by('rank')
        class_id = self.request.query_params.get('class_assigned', None)
        if class_id:
            queryset = queryset.filter(student__class_assigned_id=class_id)
        return queryset

    def _build_bulletin_elements(self, student, term, styles, doc):
        """Build ReportLab elements for a single student's bulletin."""
        elements = []

        header_data = [
            [Paragraph(SCHOOL_NAME, styles['SchoolHeader'])],
            [Paragraph(SCHOOL_ADDRESS, styles['Info'])],
        ]
        header_table = Table(header_data, colWidths=[doc.width])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (0, 0), 2),
        ]))
        elements.append(header_table)
        elements.append(HRFlowable(width="100%", thickness=2, color=PRIMARY))

        elements.append(Paragraph("BULLETIN DE NOTES", styles['BulletinTitle']))
        elements.append(Spacer(0, 4*mm))

        class_name = student.class_assigned.name if student.class_assigned else "Non assignee"
        cycle_name = ""
        if student.class_assigned and hasattr(student.class_assigned, 'cycle') and student.class_assigned.cycle:
            cycle_name = f" - {student.class_assigned.cycle.name.title()}"
        academic_year = student.academic_year or term.academic_year

        info_text = (
            f"<b>Eleve:</b> {student.last_name.upper()} {student.first_name} | "
            f"<b>Matricule:</b> {student.matricule}<br/>"
            f"<b>Classe:</b> {class_name}{cycle_name} | "
            f"<b>Trimestre:</b> {term.name} | "
            f"<b>Annee:</b> {academic_year}"
        )
        elements.append(Paragraph(info_text, styles['Info']))
        elements.append(Spacer(0, 5*mm))

        from subjects.models import Subject

        grades = Grade.objects.filter(student=student, term=term).select_related(
            'teacher_subject', 'teacher_subject__subject'
        )

        subjects = Subject.objects.none()
        if student.class_assigned and student.class_assigned.cycle:
            subjects = Subject.objects.filter(
                cycle=student.class_assigned.cycle
            ).order_by('name')

        table_data = [[
            Paragraph('N', styles['TableHeader']),
            Paragraph('Matiere', styles['TableHeader']),
            Paragraph('Note', styles['TableHeader']),
            Paragraph('Appreciation', styles['TableHeader']),
        ]]

        grade_by_ts = {g.teacher_subject_id: g for g in grades}
        ts_by_subject = {}
        for ts in TeacherSubject.objects.filter(class_assigned=student.class_assigned):
            if ts.subject_id not in ts_by_subject:
                ts_by_subject[ts.subject_id] = ts.id

        for idx, subj in enumerate(subjects, 1):
            ts_id = ts_by_subject.get(subj.id)
            grade = grade_by_ts.get(ts_id) if ts_id else None
            if grade and grade.average is not None:
                avg = grade.average
                appr = grade.appreciation
                note_str = f"{avg:.2f}"
            else:
                note_str = '--'
                appr = '--'

            table_data.append([
                Paragraph(str(idx), styles['TableCell']),
                Paragraph(subj.name, styles['TableCellLeft']),
                Paragraph(note_str, styles['TableCell']),
                Paragraph(appr, styles['TableCell']),
            ])

        table_data.append([
            Paragraph('', styles['TableCell']),
            Paragraph('<b>MOYENNE GENERALE</b>', ParagraphStyle('boldCell', parent=styles['TableCell'], textColor=PRIMARY, fontName='Helvetica-Bold')),
            Paragraph('', styles['TableCell']),
            Paragraph('', styles['TableCell']),
        ])

        grade_table = Table(table_data, colWidths=[
            10*mm, 70*mm, 25*mm, 35*mm,
        ])

        style_cmds = [
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#e5e7eb')),
            ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), HEADER_FG),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]

        for i in range(1, len(table_data) - 1):
            if i % 2 == 0:
                style_cmds.append(('BACKGROUND', (0, i), (-1, i), ROW_ALT))

        style_cmds.append(('BACKGROUND', (0, len(table_data) - 1), (-1, len(table_data) - 1), LIGHT))
        style_cmds.append(('LINEBEFORE', (0, len(table_data) - 1), (-1, len(table_data) - 1), 1.5, PRIMARY))

        table_table_style = TableStyle(style_cmds)
        grade_table.setStyle(table_table_style)
        elements.append(grade_table)
        elements.append(Spacer(0, 8*mm))

        average_obj = StudentAverage.objects.filter(student=student, term=term).first()
        if average_obj:
            gen_avg = f"{float(average_obj.average):.2f}"
            rank = f"{average_obj.rank}"
            appr = get_appreciation(average_obj.average)
        else:
            gen_avg = "Non calculee"
            rank = "--"
            appr = "--"

        summary_data = [
            [Paragraph('<b>Moyenne Generale</b>', styles['Bold']), Paragraph(gen_avg + '/20', styles['Bold']),
             Paragraph('<b>Rang</b>', styles['Bold']), Paragraph(f"{rank}/{len(grades) if grades.exists() else 1}", styles['Bold']),
             Paragraph('<b>Appreciation</b>', styles['Bold']), Paragraph(appr, styles['Bold'])],
        ]
        summary_table = Table(summary_data, colWidths=[40*mm, 25*mm, 20*mm, 25*mm, 30*mm, 30*mm])
        summary_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#e5e7eb')),
            ('BACKGROUND', (0, 0), (-1, -1), LIGHT),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(0, 15*mm))

        sig_data = [
            [
                Paragraph("Le Directeur", styles['Signature']),
                Paragraph("L'Enseignant", styles['Signature']),
                Paragraph("Le Parent/Tuteur", styles['Signature']),
            ],
            [
                Paragraph("<br/><br/>___________________", styles['Signature']),
                Paragraph("<br/><br/>___________________", styles['Signature']),
                Paragraph("<br/><br/>___________________", styles['Signature']),
            ],
        ]
        sig_table = Table(sig_data, colWidths=[55*mm, 55*mm, 55*mm])
        sig_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(sig_table)
        elements.append(Spacer(0, 5*mm))
        footer = Paragraph(
            f"<i>EduManager - Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')}</i>",
            ParagraphStyle('footer', parent=styles['Info'], alignment=TA_CENTER, fontSize=7, textColor=HexColor('#9ca3af'))
        )
        elements.append(footer)

        return elements

    @action(detail=False, methods=['get'])
    def bulletin_pdf(self, request):
        student_id = request.query_params.get('student_id')
        term_id = request.query_params.get('term_id')

        if not student_id or not term_id:
            return Response({'error': 'student_id and term_id are required'}, status=400)

        try:
            student = Student.objects.get(id=student_id)
            term = Term.objects.get(id=term_id)
        except (Student.DoesNotExist, Term.DoesNotExist):
            return Response({'error': 'Eleve ou trimestre introuvable'}, status=404)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=15*mm,
            rightMargin=15*mm,
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            name='SchoolHeader',
            fontSize=14,
            leading=18,
            alignment=TA_CENTER,
            textColor=PRIMARY,
            fontName='Helvetica-Bold',
        ))
        styles.add(ParagraphStyle(
            name='BulletinTitle',
            fontSize=12,
            leading=16,
            alignment=TA_CENTER,
            textColor=DARK,
            fontName='Helvetica-Bold',
        ))
        styles.add(ParagraphStyle(
            name='Info',
            fontSize=9,
            leading=13,
            textColor=GRAY,
            fontName='Helvetica',
        ))
        styles.add(ParagraphStyle(
            name='TableHeader',
            fontSize=8,
            leading=12,
            textColor=HEADER_FG,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
        ))
        styles.add(ParagraphStyle(
            name='TableCell',
            fontSize=8,
            leading=12,
            textColor=DARK,
            fontName='Helvetica',
            alignment=TA_CENTER,
        ))
        styles.add(ParagraphStyle(
            name='TableCellLeft',
            fontSize=8,
            leading=12,
            textColor=DARK,
            fontName='Helvetica',
            alignment=TA_LEFT,
        ))
        styles.add(ParagraphStyle(
            name='Bold',
            fontSize=9,
            leading=13,
            textColor=DARK,
            fontName='Helvetica-Bold',
        ))
        styles.add(ParagraphStyle(
            name='Signature',
            fontSize=8,
            leading=12,
            textColor=GRAY,
            fontName='Helvetica',
            alignment=TA_CENTER,
        ))

        elements = self._build_bulletin_elements(student, term, styles, doc)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"bulletin_{student.matricule}_{term.name}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'])
    def bulletins_batch_pdf(self, request):
        class_id = request.query_params.get('class_id')
        term_id = request.query_params.get('term_id')

        if not class_id or not term_id:
            return Response({'error': 'class_id and term_id are required'}, status=400)

        try:
            term = Term.objects.get(id=term_id)
            from classes.models import Class
            cls = Class.objects.get(id=class_id)
        except (Term.DoesNotExist, Class.DoesNotExist):
            return Response({'error': 'Classe ou trimestre introuvable'}, status=404)

        students = Student.objects.filter(class_assigned_id=class_id).order_by('last_name', 'first_name')
        if not students.exists():
            return Response({'error': 'Aucun eleve dans cette classe'}, status=404)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=15*mm,
            rightMargin=15*mm,
        )

        styles = getSampleStyleSheet()
        style_configs = [
            ('SchoolHeader', 14, 18, PRIMARY, 'Helvetica-Bold', TA_CENTER),
            ('BulletinTitle', 12, 16, DARK, 'Helvetica-Bold', TA_CENTER),
            ('Info', 9, 13, GRAY, 'Helvetica', TA_LEFT),
            ('TableHeader', 8, 12, HEADER_FG, 'Helvetica-Bold', TA_CENTER),
            ('TableCell', 8, 12, DARK, 'Helvetica', TA_CENTER),
            ('TableCellLeft', 8, 12, DARK, 'Helvetica', TA_LEFT),
            ('Bold', 9, 13, DARK, 'Helvetica-Bold', TA_CENTER),
            ('Signature', 8, 12, GRAY, 'Helvetica', TA_CENTER),
        ]
        for name, size, leading, color, font, align in style_configs:
            styles.add(ParagraphStyle(name=name, fontSize=size, leading=leading,
                textColor=color, fontName=font, alignment=align))

        all_elements = []
        for i, student in enumerate(students):
            if i > 0:
                all_elements.append(PageBreak())
            student_elements = self._build_bulletin_elements(student, term, styles, doc)
            all_elements.extend(student_elements)

        doc.build(all_elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"bulletins_{cls.display_name or cls.name}_{term.name}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'])
    def student_results(self, request):
        class_id = request.query_params.get('class_id')
        term_id = request.query_params.get('term_id')

        if not class_id or not term_id:
            return Response({'error': 'class_id and term_id are required'}, status=400)

        try:
            term = Term.objects.get(id=term_id)
            from classes.models import Class
            cls = Class.objects.get(id=class_id)
        except (Term.DoesNotExist, Class.DoesNotExist):
            return Response({'error': 'Classe ou trimestre introuvable'}, status=404)

        cycle_name = cls.cycle.name if cls.cycle else 'college'
        max_score = 10 if cycle_name == 'primaire' else 20
        passing_score = max_score / 2

        averages = StudentAverage.objects.filter(
            term_id=term_id,
            student__class_assigned_id=class_id,
        ).select_related('student', 'student__class_assigned', 'student__class_assigned__cycle').order_by('-average')

        results = []
        for avg in averages:
            s = avg.student
            a = float(avg.average) if avg.average is not None else 0
            mention = get_appreciation(a)
            if cycle_name == 'primaire':
                decision = 'Admis' if a >= passing_score else 'Redouble'
            else:
                decision = 'Admis' if a >= passing_score else 'Redouble'

            photo_url = None
            if s.photo:
                try:
                    photo_url = request.build_absolute_uri(s.photo.url)
                except:
                    photo_url = None

            results.append({
                'student_id': s.id,
                'matricule': s.matricule,
                'first_name': s.first_name,
                'last_name': s.last_name,
                'photo_url': photo_url,
                'class_name': cls.display_name,
                'cycle': cycle_name,
                'specialty': cls.get_specialty_display() if cls.specialty and cls.specialty != 'none' else '',
                'average': round(a, 2),
                'rank': avg.rank,
                'mention': mention,
                'decision': decision,
            })

        total = len(results)
        admis = sum(1 for r in results if r['decision'] == 'Admis')
        echoues = total - admis
        success_rate = round((admis / total * 100), 1) if total > 0 else 0
        best = results[0] if results else None
        last = results[-1] if results else None

        return Response({
            'count': total,
            'results': results,
            'stats': {
                'total': total,
                'admis': admis,
                'echoues': echoues,
                'success_rate': success_rate,
                'best_student': f"{best['first_name']} {best['last_name']}" if best else None,
                'best_avg': best['average'] if best else None,
                'last_student': f"{last['first_name']} {last['last_name']}" if last else None,
                'last_avg': last['average'] if last else None,
            },
            'class_info': {
                'name': cls.display_name,
                'cycle': cycle_name,
                'specialty': cls.get_specialty_display() if cls.specialty and cls.specialty != 'none' else '',
                'academic_year': cls.academic_year,
            },
            'term': term.name,
        })

    @action(detail=False, methods=['get'])
    def results_pdf(self, request):
        class_id = request.query_params.get('class_id')
        term_id = request.query_params.get('term_id')
        list_type = request.query_params.get('type', 'all')

        if not class_id or not term_id:
            return Response({'error': 'class_id and term_id are required'}, status=400)

        try:
            term = Term.objects.get(id=term_id)
            from classes.models import Class
            cls = Class.objects.get(id=class_id)
        except (Term.DoesNotExist, Class.DoesNotExist):
            return Response({'error': 'Classe ou trimestre introuvable'}, status=404)

        from school.models import SchoolInfo
        school = SchoolInfo.objects.first()

        cycle_name = cls.cycle.name if cls.cycle else 'college'
        max_score = 10 if cycle_name == 'primaire' else 20
        passing_score = max_score / 2

        averages = StudentAverage.objects.filter(
            term_id=term_id,
            student__class_assigned_id=class_id,
        ).select_related('student').order_by('-average')

        results = []
        for avg in averages:
            s = avg.student
            a = float(avg.average) if avg.average is not None else 0
            decision = 'Admis' if a >= passing_score else 'Redouble'
            if list_type == 'admis' and decision != 'Admis':
                continue
            if list_type == 'echoues' and decision == 'Admis':
                continue
            mention = get_appreciation(a)
            results.append({
                'student': s,
                'average': a,
                'rank': avg.rank,
                'decision': decision,
                'mention': mention,
            })

        if not results:
            return Response({'error': 'Aucun resultat pour cette liste'}, status=404)

        MENTION_ORDER = {'Tres bien': 0, 'Bien': 1, 'Assez bien': 2, 'Passable': 3, 'Insuffisant': 4}
        results.sort(key=lambda r: (MENTION_ORDER.get(r['mention'], 99), -r['average']))

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            topMargin=15*mm, bottomMargin=20*mm,
            leftMargin=15*mm, rightMargin=15*mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title2', fontSize=14, leading=18, textColor=PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)
        subtitle_style = ParagraphStyle('Sub', fontSize=10, leading=14, textColor=DARK, fontName='Helvetica', alignment=TA_CENTER)
        th_style = ParagraphStyle('TH', fontSize=8, leading=12, textColor=white, fontName='Helvetica-Bold', alignment=TA_CENTER)
        td_style = ParagraphStyle('TD', fontSize=8, leading=12, textColor=DARK, fontName='Helvetica', alignment=TA_CENTER)
        td_left = ParagraphStyle('TDL', fontSize=8, leading=12, textColor=DARK, fontName='Helvetica', alignment=TA_LEFT)
        stat_style = ParagraphStyle('Stat', fontSize=10, leading=14, textColor=GRAY, fontName='Helvetica', alignment=TA_CENTER)
        signature_style = ParagraphStyle('Sig', fontSize=8, leading=12, textColor=GRAY, fontName='Helvetica', alignment=TA_CENTER)

        elements = []

        if school and school.logo:
            try:
                from reportlab.platypus import Image as RLImage
                logo = RLImage(school.logo.path, width=40*mm, height=15*mm)
                logo_table = Table([[logo]], colWidths=[doc.width])
                logo_table.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER')]))
                elements.append(logo_table)
            except:
                pass

        school_name = school.name if school else SCHOOL_NAME
        elements.append(Paragraph(school_name, title_style))
        elements.append(Paragraph(f"Annee scolaire: {cls.academic_year}", subtitle_style))
        elements.append(Spacer(0, 2*mm))

        title_text = "LISTE DES ELEVES ADMIS" if list_type != 'echoues' else "LISTE DES ELEVES ECHOUS"
        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(0, 3*mm))

        info_text = f"Classe: {cls.display_name} | Cycle: {cls.cycle.get_name_display() if cls.cycle else ''} | Trimestre: {term.name}"
        if cls.specialty and cls.specialty != 'none':
            info_text += f" | Serie: {cls.get_specialty_display()}"
        elements.append(Paragraph(info_text, subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=PRIMARY))
        elements.append(Spacer(0, 4*mm))

        table_data = [[
            Paragraph('N°', th_style),
            Paragraph('Matricule', th_style),
            Paragraph('Nom & Prenom', th_style),
            Paragraph(f'Moy./{max_score}', th_style),
            Paragraph('Rang', th_style),
            Paragraph('Decision', th_style),
        ]]

        mention_rows = []
        current_mention = None
        idx = 0
        for r in results:
            if r['mention'] != current_mention:
                current_mention = r['mention']
                mention_label = {'Tres bien': 'Très Bien', 'Bien': 'Bien', 'Assez bien': 'Assez Bien', 'Passable': 'Passable', 'Insuffisant': 'Insuffisant'}.get(current_mention, current_mention)
                mention_color = {'Tres bien': '#1e40af', 'Bien': '#16a34a', 'Assez bien': '#ca8a04', 'Passable': '#ea580c', 'Insuffisant': '#dc2626'}.get(current_mention, '#6b7280')
                mention_rows.append(len(table_data))
                table_data.append([
                    Paragraph('', th_style),
                    Paragraph('', th_style),
                    Paragraph(f"<b>{mention_label}</b>", ParagraphStyle('mention', parent=td_left, textColor=HexColor(mention_color), fontName='Helvetica-Bold')),
                    Paragraph('', th_style),
                    Paragraph('', th_style),
                    Paragraph('', th_style),
                ])
            idx += 1
            dec_color = GREEN if r['decision'] == 'Admis' else ORANGE
            table_data.append([
                Paragraph(str(idx), td_style),
                Paragraph(r['student'].matricule, td_style),
                Paragraph(f"{r['student'].last_name} {r['student'].first_name}", td_left),
                Paragraph(f"{r['average']:.2f}", td_style),
                Paragraph(str(r['rank'] or '--'), td_style),
                Paragraph(f"<b>{r['decision']}</b>", ParagraphStyle('dec', parent=td_style, textColor=dec_color)),
            ])

        col_w = [10*mm, 28*mm, 65*mm, 22*mm, 15*mm, 25*mm]
        grade_table = Table(table_data, colWidths=col_w, repeatRows=1)
        style_cmds = [
            ('GRID', (0,0), (-1,-1), 0.5, HexColor('#d1d5db')),
            ('BACKGROUND', (0,0), (-1,0), PRIMARY),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
        ]
        for i in range(1, len(table_data)):
            if i in mention_rows:
                style_cmds.append(('BACKGROUND', (0,i), (-1,i), HexColor('#f1f5f9')))
                style_cmds.append(('SPAN', (0,i), (-1,i)))
            elif (i - sum(1 for m in mention_rows if m < i)) % 2 == 1:
                style_cmds.append(('BACKGROUND', (0,i), (-1,i), ROW_ALT))
        grade_table.setStyle(TableStyle(style_cmds))
        elements.append(grade_table)
        elements.append(Spacer(0, 6*mm))

        total = len(results)
        admis = sum(1 for r in results if r['decision'] == 'Admis')
        echoues = total - admis
        rate = round((admis / total * 100), 1) if total > 0 else 0

        stats_data = [
            [Paragraph(f"Total: {total}", stat_style)],
            [Paragraph(f"Admis: {admis}", ParagraphStyle('sg', parent=stat_style, textColor=GREEN))],
            [Paragraph(f"Echoues: {echoues}", ParagraphStyle('sr', parent=stat_style, textColor=ORANGE))],
            [Paragraph(f"Taux reussite: {rate}%", stat_style)],
        ]
        stats_table = Table(stats_data, colWidths=[doc.width])
        stats_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(0, 8*mm))

        sig_data = [
            [Paragraph("Le Directeur", signature_style), Paragraph("Le Parent/Tuteur", signature_style)],
            [Paragraph("<br/><br/>___________________", signature_style), Paragraph("<br/><br/>___________________", signature_style)],
        ]
        sig_table = Table(sig_data, colWidths=[80*mm, 80*mm])
        sig_table.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER')]))
        elements.append(sig_table)
        elements.append(Spacer(0, 5*mm))
        elements.append(Paragraph(
            f"<i>Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')} - EduManager</i>",
            ParagraphStyle('footer', parent=subtitle_style, fontSize=7, textColor=HexColor('#9ca3af'))
        ))

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        type_label = "admis" if list_type != 'echoues' else "echoues"
        filename = f"{type_label}_{cls.display_name}_{term.name}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'])
    def results_excel(self, request):
        class_id = request.query_params.get('class_id')
        term_id = request.query_params.get('term_id')
        list_type = request.query_params.get('type', 'all')

        if not class_id or not term_id:
            return Response({'error': 'class_id and term_id are required'}, status=400)

        try:
            term = Term.objects.get(id=term_id)
            from classes.models import Class
            cls = Class.objects.get(id=class_id)
        except (Term.DoesNotExist, Class.DoesNotExist):
            return Response({'error': 'Classe ou trimestre introuvable'}, status=404)

        cycle_name = cls.cycle.name if cls.cycle else 'college'
        max_score = 10 if cycle_name == 'primaire' else 20
        passing_score = max_score / 2

        averages = StudentAverage.objects.filter(
            term_id=term_id,
            student__class_assigned_id=class_id,
        ).select_related('student').order_by('-average')

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        from school.models import SchoolInfo
        school = SchoolInfo.objects.first()
        school_name = school.name if school else SCHOOL_NAME

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultats"

        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='d1d5db'),
            right=Side(style='thin', color='d1d5db'),
            top=Side(style='thin', color='d1d5db'),
            bottom=Side(style='thin', color='d1d5db'),
        )

        ws.cell(row=1, column=1, value=f"{school_name if True else 'Ecole'} - {cls.display_name}").font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=1, value=f"Trimestre: {term.name} | Annee: {cls.academic_year}").font = Font(size=10, color='6b7280')
        ws.merge_cells('A2:F2')
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

        headers = ['N°', 'Matricule', 'Nom & Prenom', f'Moyenne/{max_score}', 'Rang', 'Decision']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        row_num = 5
        idx = 1
        for avg in averages:
            s = avg.student
            a = float(avg.average) if avg.average is not None else 0
            decision = 'Admis' if a >= passing_score else 'Redouble'
            if list_type == 'admis' and decision != 'Admis':
                continue
            if list_type == 'echoues' and decision == 'Admis':
                continue

            dec_font = Font(color='16a34a', bold=True) if decision == 'Admis' else Font(color='ea580c', bold=True)
            row_fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid') if row_num % 2 == 0 else None

            ws.cell(row=row_num, column=1, value=idx).border = thin_border
            ws.cell(row=row_num, column=2, value=s.matricule).border = thin_border
            ws.cell(row=row_num, column=3, value=f"{s.last_name} {s.first_name}").border = thin_border
            ws.cell(row=row_num, column=4, value=round(a, 2)).border = thin_border
            ws.cell(row=row_num, column=5, value=avg.rank or '').border = thin_border
            cell_dec = ws.cell(row=row_num, column=6, value=decision)
            cell_dec.font = dec_font
            cell_dec.border = thin_border
            cell_dec.alignment = Alignment(horizontal='center')

            if row_fill:
                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).fill = row_fill
            row_num += 1
            idx += 1

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 32
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 14

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        type_label = "all" if list_type == 'all' else list_type
        filename = f"resultats_{type_label}_{cls.display_name}_{term.name}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
